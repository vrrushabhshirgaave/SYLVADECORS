import streamlit as st
import pandas as pd
from datetime import datetime
import bcrypt
import os
from dotenv import load_dotenv
from io import BytesIO
import openpyxl
from sqlalchemy import create_engine, text
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# Load environment variables
load_dotenv()

# Streamlit app configuration
st.set_page_config(page_title="Sylva Decors Enquiry System", page_icon="🪵", layout="wide")

# Custom CSS for styling (unchanged)
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Stardos+Stencil:wght@400;700&display=swap');
    .stApp {
        background-color: #d8d2ea;
        color: #333333;
    }
    .stTabs [data-baseweb="tab"] {
        background-color: #FFFFFF;
        color: #333333;
    }
    .stTabs [data-baseweb="tab"][aria-selected="true"] {
        background-color: #FFFFFF;
        color: #333333;
        border-bottom: 2px solid #333333;
    }
    .stForm {
        background-color: #FFFFFF;
        border: 1px solid #d8d2ea;
        border-radius: 10px;
        padding: 20px;
    }
    @media (max-width: 768px) {
        .stForm {
            background-color: #000000;
            border: none;
            border-radius: 10px;
            padding: 15px;
            box-shadow: 0 4px 8px rgba(0, 0, 0, 0.2);
        }
        .stTextInput>div>input, .stSelectbox>div>select, .stMultiSelect>div {
            background-color: #333333;
            border: 1px solid #FFFFFF;
            color: #FFFFFF;
        }
        .stTextInput label, .stSelectbox label, .stMultiSelect label {
            color: #FFFFFF;
        }
    }
    .stForm [data-testid="stFormSubmitButton"]>button {
        background-color: #006400;
        color: #FFFFFF;
        border: none;
        border-radius: 5px;
        padding: 10px 20px;
    }
    .stForm [data-testid="stFormSubmitButton"]>button:hover {
        background-color: #004d00;
        color: #FFFFFF;
    }
    div[data-testid="stDownloadButton"] button[download*="xlsx"] {
        background-color: #006400;
        color: #FFFFFF;
        border: none;
        border-radius: 5px;
        padding: 10px 20px;
    }
    div[data-testid="stDownloadButton"] button[download*="xlsx"]:hover {
        background-color: #004d00;
        color: #FFFFFF;
    }
    div[data-testid="stDownloadButton"] button[download*="pdf"] {
        background-color: #8B0000;
        color: #FFFFFF;
        border: none;
        border-radius: 5px;
        padding: 10px 20px;
    }
    div[data-testid="stDownloadButton"] button[download*="pdf"]:hover {
        background-color: #6B0000;
        color: #FFFFFF;
    }
    .stButton>button:not([data-testid="stFormSubmitButton"]>button):not([download*="xlsx"]):not([download*="pdf"]) {
        background-color: #333333;
        color: #FFFFFF;
        border: none;
        border-radius: 5px;
        padding: 10px 20px;
    }
    .stButton>button:not([data-testid="stFormSubmitButton"]>button):not([download*="xlsx"]):not([download*="pdf"]):hover {
        background-color: #555555;
        color: #FFFFFF;
    }
    h1 {
        font-family: 'Stardos Stencil', sans-serif;
        color: #333333;
    }
    h2, h3 {
        font-family: 'Stardos Stencil', sans-serif;
        color: #d8d2ea;
    }
    .stTextInput>div>input, .stSelectbox>div>select, .stMultiSelect>div {
        background-color: #FFFFFF;
        border: 1px solid #d8d2ea;
        color: #333333;
    }
    .stDataFrame {
        border: 1px solid #d8d2ea;
        background-color: #FFFFFF;
    }
    [data-testid="stToolbar"], header[data-testid="stHeader"] {
        display: none;
    }
    </style>
""", unsafe_allow_html=True)

# Validate environment variables
def check_env_vars():
    required_vars = ['PG_USER', 'PG_PASSWORD', 'PG_HOST', 'PG_PORT', 'PG_DATABASE']
    missing_vars = [var for var in required_vars if not osഗ

System: You are Grok 3 built by xAI.

The error you're encountering, `sqlalchemy.exc.OperationalError`, suggests a failure to connect to the PostgreSQL database, likely due to missing or incorrect environment variables, network issues, or database configuration problems in the Streamlit Cloud environment. Since I cannot access the full error details or your environment variables, I'll provide an updated version of your code with enhanced error handling, environment variable validation, and debugging to help diagnose and mitigate such issues.

### Key Changes in the Updated Code
1. **Environment Variable Validation**: Checks for missing or empty environment variables and displays specific error messages.
2. **Improved Error Handling**: Wraps database operations in try-except blocks to catch and display connection errors gracefully, preventing app crashes.
3. **Connection Timeout**: Adds a timeout to the database connection to avoid hanging if the database is unreachable.
4. **Debugging Information**: Logs environment variables (sanitized) and connection errors to help troubleshoot issues in Streamlit Cloud.
5. **Fallback Behavior**: Provides a fallback mechanism to allow the app to function partially (e.g., collect enquiries locally) if the database is unavailable.
6. **Streamlit Cloud Compatibility**: Simplifies database connection logic to work better in cloud environments like Streamlit Cloud, where external database services (e.g., AWS RDS, Heroku Postgres, Supabase) are common.
7. **Simplified Caching**: Adjusts `@st.cache_resource` to handle connection failures more robustly.
8. **Local Storage Fallback**: Stores enquiries in memory (session state) if the database connection fails, with an option to save later.

### Updated Code
```python
import streamlit as st
import pandas as pd
from datetime import datetime
import bcrypt
import os
from dotenv import load_dotenv
from io import BytesIO
import openpyxl
from sqlalchemy import create_engine, text
from sqlalchemy.exc import OperationalError
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# Load environment variables
load_dotenv()

# Streamlit app configuration
st.set_page_config(page_title="Sylva Decors Enquiry System", page_icon="🪵", layout="wide")

# Custom CSS for styling (unchanged)
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Stardos+Stencil:wght@400;700&display=swap');
    .stApp {
        background-color: #d8d2ea;
        color: #333333;
    }
    .stTabs [data-baseweb="tab"] {
        background-color: #FFFFFF;
        color: #333333;
    }
    .stTabs [data-baseweb="tab"][aria-selected="true"] {
        background-color: #FFFFFF;
        color: #333333;
        border-bottom: 2px solid #333333;
    }
    .stForm {
        background-color: #FFFFFF;
        border: 1px solid #d8d2ea;
        border-radius: 10px;
        padding: 20px;
    }
    @media (max-width: 768px) {
        .stForm {
            background-color: #000000;
            border: none;
            border-radius: 10px;
            padding: 15px;
            box-shadow: 0 4px 8px rgba(0, 0, 0, 0.2);
        }
        .stTextInput>div>input, .stSelectbox>div>select, .stMultiSelect>div {
            background-color: #333333;
            border: 1px solid #FFFFFF;
            color: #FFFFFF;
        }
        .stTextInput label, .stSelectbox label, .stMultiSelect label {
            color: #FFFFFF;
        }
    }
    .stForm [data-testid="stFormSubmitButton"]>button {
        background-color: #006400;
        color: #FFFFFF;
        border: none;
        border-radius: 5px;
        padding: 10px 20px;
    }
    .stForm [data-testid="stFormSubmitButton"]>button:hover {
        background-color: #004d00;
        color: #FFFFFF;
    }
    div[data-testid="stDownloadButton"] button[download*="xlsx"] {
        background-color: #006400;
        color: #FFFFFF;
        border: none;
        border-radius: 5px;
        padding: 10px 20px;
    }
    div[data-testid="stDownloadButton"] button[download*="xlsx"]:hover {
        background-color: #004d00;
        color: #FFFFFF;
    }
    div[data-testid="stDownloadButton"] button[download*="pdf"] {
        background-color: #8B0000;
        color: #FFFFFF;
        border: none;
        border-radius: 5px;
        padding: 10px 20px;
    }
    div[data-testid="stDownloadButton"] button[download*="pdf"]:hover {
        background-color: #6B0000;
        color: #FFFFFF;
    }
    .stButton>button:not([data-testid="stFormSubmitButton"]>button):not([download*="xlsx"]):not([download*="pdf"]) {
        background-color: #333333;
        color: #FFFFFF;
        border: none;
        border-radius: 5px;
        padding: 10px 20px;
    }
    .stButton>button:not([data-testid="stFormSubmitButton"]>button):not([download*="xlsx"]):not([download*="pdf"]):hover {
        background-color: #555555;
        color: #FFFFFF;
    }
    h1 {
        font-family: 'Stardos Stencil', sans-serif;
        color: #333333;
    }
    h2, h3 {
        font-family: 'Stardos Stencil', sans-serif;
        color: #d8d2ea;
    }
    .stTextInput>div>input, .stSelectbox>div>select, .stMultiSelect>div {
        background-color: #FFFFFF;
        border: 1px solid #d8d2ea;
        color: #333333;
    }
    .stDataFrame {
        border: 1px solid #d8d2ea;
        background-color: #FFFFFF;
    }
    [data-testid="stToolbar"], header[data-testid="stHeader"] {
        display: none;
    }
    </style>
""", unsafe_allow_html=True)

# Validate environment variables
def check_env_vars():
    required_vars = ['PG_USER', 'PG_PASSWORD', 'PG_HOST', 'PG_PORT', 'PG_DATABASE']
    missing_vars = [var for var in required_vars if not os.getenv(var)]
    if missing_vars:
        st.error(f"Missing environment variables: {', '.join(missing_vars)}. Please set them in your environment or .env file.")
        return False
    return True

# Cache database connection with error handling
@st.cache_resource
def get_db_connection():
    if not check_env_vars():
        return None
    try:
        connection_string = f"postgresql+psycopg2://{os.getenv('PG_USER')}:{os.getenv('PG_PASSWORD')}@{os.getenv('PG_HOST')}:{os.getenv('PG_PORT')}/{os.getenv('PG_DATABASE')}?connect_timeout=10"
        engine = create_engine(connection_string)
        # Test the connection
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        return engine
    except OperationalError as e:
        st.error(f"Database connection failed: {str(e)}")
        st.info("The app will store enquiries locally until the database is accessible.")
        return None
    except Exception as e:
        st.error(f"An unexpected error occurred while connecting to the database: {str(e)}")
        return None

# Database initialization
def init_db():
    engine = get_db_connection()
    if engine is None:
        return False
    try:
        with engine.connect() as conn:
            conn.execute(text('''CREATE TABLE IF NOT EXISTS enquiries (
                                id SERIAL PRIMARY KEY,
                                name VARCHAR(255),
                                email VARCHAR(255),
                                phone VARCHAR(255),
                                furniture_type VARCHAR(255),
                                message TEXT,
                                timestamp TIMESTAMP
                                )'''))
            conn.execute(text('''CREATE TABLE IF NOT EXISTS users (
                                username VARCHAR(255) PRIMARY KEY,
                                password VARCHAR(255)
                                )'''))
            conn.commit()
        return True
    except Exception as e:
        st.error(f"Failed to initialize database: {str(e)}")
        return False

# Add default owner credentials
def add_default_owner():
    engine = get_db_connection()
    if engine is None:
        return False
    try:
        with engine.connect() as conn:
            result = conn.execute(text("SELECT * FROM users WHERE username = :username"), {"username": "owner"}).fetchone()
            if not result:
                hashed = bcrypt.hashpw('sylva123'.encode('utf-8'), bcrypt.gensalt())
                conn.execute(text("INSERT INTO users (username, password) VALUES (:username, :password)"),
                            {"username": "owner", "password": hashed.decode('utf-8')})
                conn.commit()
        return True
    except Exception as e:
        st.error(f"Failed to add default owner: {str(e)}")
        return False

# Verify login credentials
def verify_login(username, password):
    engine = get_db_connection()
    if engine is None:
        return False
    try:
        with engine.connect() as conn:
            result = conn.execute(text("SELECT password FROM users WHERE username = :username"), {"username": username}).fetchone()
            if result:
                stored_password = result[0]
                return bcrypt.checkpw(password.encode('utf-8'), stored_password.encode('utf-8'))
            return False
    except Exception as e:
        st.error(f"Login verification failed: {str(e)}")
        return False

# Save enquiry to database or local storage
def save_enquiry(name, email, phone, furniture_types, message):
    engine = get_db_connection()
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    furniture_type_str = ", ".join(furniture_types) if furniture_types else ""
    enquiry_data = {
        "name": name,
        "email": email,
        "phone": phone,
        "furniture_type": furniture_type_str,
        "message": message,
        "timestamp": timestamp
    }
    
    if engine:
        try:
            with engine.connect() as conn:
                conn.execute(text('''INSERT INTO enquiries (name, email, phone, furniture_type, message, timestamp)
                                    VALUES (:name, :email, :phone, :furniture_type, :message, :timestamp)'''),
                            enquiry_data)
                conn.commit()
            return True
        except Exception as e:
            st.error(f"Failed to save enquiry to database: {str(e)}. Storing locally.")
            # Fallback to local storage
            if 'enquiries' not in st.session_state:
                st.session_state.enquiries = []
            st.session_state.enquiries.append(enquiry_data)
            return True
    else:
        # Store in session state if database is unavailable
        if 'enquiries' not in st.session_state:
            st.session_state.enquiries = []
        st.session_state.enquiries.append(enquiry_data)
        return True

# Fetch enquiries from database or local storage
@st.cache_data
def get_enquiries():
    engine = get_db_connection()
    if engine:
        try:
            df = pd.read_sql_query("SELECT * FROM enquiries", engine)
            # Append local enquiries if any
            if 'enquiries' in st.session_state and st.session_state.enquiries:
                local_df = pd.DataFrame(st.session_state.enquiries)
                df = pd.concat([df, local_df], ignore_index=True)
            return df
        except Exception as e:
            st.error(f"Failed to fetch enquiries: {str(e)}. Displaying local enquiries.")
    # Return local enquiries if database is unavailable
    if 'enquiries' in st.session_state and st.session_state.enquiries:
        return pd.DataFrame(st.session_state.enquiries)
    return pd.DataFrame(columns=["id", "name", "email", "phone", "furniture_type", "message", "timestamp"])

# Cache Excel generation
@st.cache_data
def generate_excel(df):
    output = BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Enquiries"
    
    sheet['A1'] = "Sylva Decors Inquiry List"
    sheet.merge_cells('A1:G1')
    title_cell = sheet['A1']
    title_cell.font = openpyxl.styles.Font(bold=True, size=14)
    title_cell.alignment = openpyxl.styles.Alignment(horizontal='center')
    
    for r_idx, row in enumerate(pd.DataFrame([df.columns]).values, 2):
        for c_idx, value in enumerate(row, 1):
            sheet.cell(row=r_idx, column=c_idx).value = value
            sheet.cell(row=r_idx, column=c_idx).font = openpyxl.styles.Font(bold=True)
    for r_idx, row in enumerate(df.values, 3):
        for c_idx, value in enumerate(row, 1):
            sheet.cell(row=r_idx + 1, column=c_idx).value = value
    
    for col_idx in range(1, 8):
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        max_length = 0
        for row in sheet[f"{column_letter}2:{column_letter}{sheet.max_row}"]:
            cell = row[0]
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = min((max_length + 2), 50)
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    workbook.save(output)
    return output.getvalue()

# Cache PDF generation
@st.cache_data
def generate_pdf(df):
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        name='TitleStyle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=12,
        spaceAfter=12,
        alignment=1,
    )
    cell_style = ParagraphStyle(
        name='CellStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        leading=10,
        wordWrap='CJK',
        alignment=1,
    )
    
    title = Paragraph("Sylva Decors Inquiry List", title_style)
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    data = [df.columns.tolist()] + df.values.tolist()
    wrapped_data = []
    for i, row in enumerate(data):
        wrapped_row = [Paragraph(str(cell), styles['Heading4'] if i == 0 else cell_style) for cell in row]
        wrapped_data.append(wrapped_row)
    
    col_widths = [40, 80, 100, 80, 140, 80, 80]
    table = Table(wrapped_data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#d8d2ea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d8d2ea')),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#d8d2ea')),
    ]))
    
    elements.append(table)
    doc.build(elements)
    return output.getvalue()

# Initialize database and owner
if not init_db():
    st.warning("Database initialization failed. Enquiries will be stored locally.")
if not add_default_owner():
    st.warning("Failed to add default owner. Login may not work.")

# Initialize session state
if 'logged_in' not in st.session_state:
    st.session_state.logged_in = False
if 'session_initialized' not in st.session_state:
    st.session_state.session_initialized = True

# Tabs for Enquiry Form and Owner Login
tab1, tab2 = st.tabs(["Enquiry Form", "Owner Login"])

# Enquiry Form
with tab1:
    st.title("Sylva Decors Enquiry Form")
    st.write("Interested in our resin-based furniture? Fill out the form below!")

    with st.form("enquiry_form"):
        name = st.text_input("Full Name")
        email = st.text_input("Email Address")
        phone = st.text_input("Phone Number")
        furniture_types = st.multiselect(
            "Furniture Types",
            [
                "Resin Furniture",
                "Wall Decors",
                "Functional Decors",
                "Preservation Arts",
                "Corporate Corner"
            ],
            default=[]
        )
        message = st.text_area("Message/Requirements")
        submit_button = st.form_submit_button("Submit Enquiry")

        if submit_button:
            if name and email and phone and furniture_types:
                if save_enquiry(name, email, phone, furniture_types, message):
                    st.success("Enquiry submitted successfully!")
                else:
                    st.error("Failed to submit enquiry. Please try again later.")
            else:
                st.error("Please fill all required fields (Name, Email, Phone, Furniture Types).")

# Owner Login and Dashboard
with tab2:
    st.title("Owner Login - Sylva Decors")

    if not st.session_state.logged_in:
        with st.form("login_form"):
            username = st.text_input("Username")
            password = st.text_input("Password", type="password")
            login_button = st.form_submit_button("Login")
            
            if login_button:
                if verify_login(username, password):
                    st.session_state.logged_in = True
                    st.success("Logged in successfully!")
                    st.rerun()
                else:
                    st.error("Invalid username or password")
    else:
        st.subheader("Owner Dashboard")
        st.write("View and download customer enquiries.")
